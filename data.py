import os
import openpyxl as opxl
import requests
import contextlib

from db import create_bsu_table, create_savings_table, create_savings_limit_table, create_retirement_table, create_usage_and_salary_table, insert_bsu, insert_savings_account, insert_savings_limit_account, insert_retirement, insert_usage_and_salary, get_bsu_banks, get_savings_banks, get_savings_limit_banks, get_retirement_banks, get_usage_and_salary_banks

from settings import access_spreadsheet, bsu_count, url_bsu_regions, url_bsu_country, url_savings_acc_nolimit, url_savings_acc_limit_34less, url_savings_acc_limit_34more, url_retirement_savings, url_usage_and_salary, check_codec, time_now


# Method for downloading finansportalen content
def download_content(url, filename):
    if not os.path.exists('download'):
        os.makedirs('download')

    response = requests.get(url)
    with open(os.path.join('download', filename), 'wb') as f:
        f.write(response.content)


# Method for updating Google form data
def update_form(banks, sheet_number):
    # Initialize spreadsheet and worksheet
    sheet = access_spreadsheet()
    bank_sheet = sheet.get_worksheet(sheet_number)
    row = 3
    after_row = (len(banks) + 3)

    # Insert all data into sheet
    print('Adding banks to form.')
    for x in range(len(banks)):
        bank_sheet.update_acell('A%s' % (row), banks[x])
        row += 1

    # Deleting all banks that are no longer represented
    print('Deleting old banks from form.')
    while True:
        next_cell = bank_sheet.acell('A%s' % (after_row))

        if not next_cell.value == '':
            bank_sheet.update_acell('A%s' % (after_row), '')
        else:
            break

        after_row += 1


# Method for extracting column data from file
def extract_xlsx_data(file):
    data = []

    # Load file
    wb = opxl.load_workbook('download/' + file)

    # Open worksheet
    work_sheet = time_now('date.month.year')
    ws = wb.get_sheet_by_name(work_sheet)

    # Read specific columns row by row contents from file
    for row in range(2, ws.max_row + 1):
        data_item = []
        for column in "BLOMPDEH":
            cell_name = "{}{}".format(column, row)
            cell_data = ws[cell_name].value
            data_item.append(cell_data)

        # Only list the best savings account from each bank
        exist = True
        for y in range(len(data)):
            if data_item[2] in data[y][2]:
                exist = False

        if exist:
            data.append(data_item)

    # Delete temporary file downloaded
    #with contextlib.suppress(FileNotFoundError):
    #    os.remove('download/' + file)

    return data


# Method for handling and storing extracted data
def handle_and_store_data(data, db_table, limit_age):
    product_id = ''
    bank_id = ''
    bank_name = ''
    bank_url = ''
    bank_region = ''
    bank_account_name = ''
    publication_date = ''
    interest_rate = ''

    # Loop through rows in Excel file
    for x in range(len(data)):
        product_id = str(int(data[x][0]))
        bank_id = str(int(data[x][1]))

        bank_name = data[x][2].encode('ISO-8859-1', 'ignore')
        bank_name = check_codec(bank_name).decode('ISO-8859-1')

        bank_url = data[x][3]
        if 'http:' in bank_url:
            bank_url = bank_url.replace('http:', 'https:')  # For security reasons

        bank_region = data[x][4]
        bank_account_name = data[x][5]
        publication_date = data[x][6]
        interest_rate = float("%.2f" % data[x][7])


        if db_table == 'bsu':
            # Insert bsu data into database
            insert_bsu(product_id, bank_id, bank_name, bank_url, bank_region, bank_account_name, publication_date,
                       interest_rate)
        elif db_table == 'savings_account':
            # Insert bsu data into database
            insert_savings_account(product_id, bank_id, bank_name, bank_url, bank_region, bank_account_name,
                                   publication_date, interest_rate)
        elif db_table == 'savings_account_limit':
            insert_savings_limit_account(product_id, bank_id, bank_name, bank_url, bank_region, bank_account_name,
                                         publication_date, interest_rate, limit_age)
        elif db_table == 'retirement':
            insert_retirement(product_id, bank_id, bank_name, bank_url, bank_region, bank_account_name, publication_date,
                              interest_rate)
        elif db_table == 'usage_salary':
            insert_usage_and_salary(product_id, bank_id, bank_name, bank_url, bank_region, bank_account_name, publication_date,
                                    interest_rate)


# Method for handling BSU data
def get_bsu_data():
    bsu_files = ['bsu_regions.xlsx', 'bsu_country.xlsx']
    bsu_urls = [url_bsu_regions, url_bsu_country]

    # Reset BSU database table
    create_bsu_table()

    # Looping through the data content
    for x in range(bsu_count):
        bsu_data = []

        # Setting local bsu variables
        bsu_file = bsu_files[x]
        bsu_url = bsu_urls[x]

        # Download data content
        #download_content(bsu_url, bsu_file)

        # Extract data from xlsx file
        bsu_data = extract_xlsx_data(bsu_file)

        # Handle and store data
        handle_and_store_data(bsu_data, 'bsu', '')

    # Get BSU bank names from BSU database
    bsu_banks = get_bsu_banks()

    print('Start Google Spreadsheet handling for BSU data.')
    #update_form(bsu_banks, 2)


# Method for handling savings account (with no limit) data
def get_savings_acc_nolimit_data():
    savings_file = 'savings_account_nolimit.xlsx'
    savings_url = url_savings_acc_nolimit

    # Reset savings account database table
    create_savings_table()

    # Download data content
    #download_content(savings_url, savings_file)

    # Extract data from xlsx file
    savings_data = extract_xlsx_data(savings_file)

    # Handle and store data
    handle_and_store_data(savings_data, 'savings_account', '')

    # Get savings bank names from savings account database
    savings_banks = get_savings_banks()

    print('Start Google Spreadsheet handling for savings data.')
    #update_form(savings_banks, 3)


# Method for handling savings account (with limit) data
def get_savings_acc_limit_data():
    savings_files = ['savings_account_limit_34h.xlsx', 'savings_account_limit_34l.xlsx']
    savings_urls = [url_savings_acc_limit_34more, url_savings_acc_limit_34less]
    limit_age = [0, 1]

    # Reset savings limit account database table
    create_savings_limit_table()

    # Looping through the data content
    for x in range(len(savings_files)):
        savings_limit_data = []

        # Setting local savings variables
        savings_file = savings_files[x]
        savings_url = savings_urls[x]

        # Download data content
        #download_content(savings_url, savings_file)

        # Extract data from xlsx file
        savings_limit_data = extract_xlsx_data(savings_file)

        # Handle and store data
        handle_and_store_data(savings_limit_data, 'savings_account_limit', limit_age[x])

    # Get BSU bank names from BSU database
    savings_limit_banks = get_savings_limit_banks()

    print('Start Google Spreadsheet handling for savings limit banks data.')
    #update_form(savings_limit_banks, 4)


# Method for handling retirement data
def get_retirement_data():
    retirement_file = 'retirement_saving.xlsx'
    retirement_url = url_retirement_savings

    # Reset savings account database table
    create_retirement_table()

    # Download data content
    #download_content(retirement_url, retirement_file)

    # Extract data from xlsx file
    retirement_data = extract_xlsx_data(retirement_file)

    # Handle and store data
    handle_and_store_data(retirement_data, 'retirement', '')

    # Get savings bank names from savings account database
    retirement_banks = get_retirement_banks()

    print('Start Google Spreadsheet handling for retirement data.')
    #update_form(retirement_banks, 5)


# Method for handling usage and salary data
def get_usage_and_salary_data():
    usage_and_salary_file = 'usage_and_salary.xlsx'
    usage_and_salary_url = url_usage_and_salary

    # Reset savings account database table
    create_usage_and_salary_table()

    # Download data content
    #download_content(usage_and_salary_url, usage_and_salary_file)

    # Extract data from xlsx file
    usage_and_salary_data = extract_xlsx_data(usage_and_salary_file)

    # Handle and store data
    handle_and_store_data(usage_and_salary_data, 'usage_salary', '')

    # Get savings bank names from savings account database
    usage_and_salary_banks = get_usage_and_salary_banks()

    print('Start Google Spreadsheet handling for usage and salary data.')
    #update_form(usage_and_salary_banks, 6)
